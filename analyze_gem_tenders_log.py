#!/usr/bin/env python3
"""
Script to analyze gem_tenders.log file per-date
Provides summary analysis including:
- Organization ID
- Number of tenders downloaded
- Number of tenders selected after keyword filter
- Number of tenders analyzed
- Errors (warnings ignored)

Features:
- Creates separate Excel sheets for each date
- Auto-purges sheets older than 15 days
"""

import re
import os
import glob
from collections import defaultdict
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def parse_log_file(log_file_path):
    """
    Parse the gem_tenders.log file and extract relevant metrics, grouped by date.

    Returns:
        dict: Analysis results grouped by date (YYYY-MM-DD format)
    """

    analysis = defaultdict(lambda: defaultdict(lambda: {
        'org_id': None,
        'downloaded_tenders': 0,
        'filtered_out_tenders': 0,
        'selected_tenders': 0,
        'analyzed_tenders': 0,
        'errors': [],
        'session_start': None,
        'session_end': None
    }))

    session_counter = 0
    current_session = None

    try:
        with open(log_file_path, 'r', encoding='utf-8') as f:
            for line_num, line in enumerate(f, 1):
                try:
                    # Extract timestamp
                    timestamp_match = re.match(r'(\d{4}-\d{2}-\d{2}) (\d{2}:\d{2}:\d{2})', line)
                    if not timestamp_match:
                        continue

                    date = timestamp_match.group(1)
                    time = timestamp_match.group(2)
                    timestamp = f"{date} {time}"

                    # Check if this is a new session (detect session start patterns)
                    if 'Running' in line and 'gem_nlp_api' in line and 'interactive mode' in line:
                        session_counter += 1
                        current_session = session_counter

                    if current_session is None:
                        current_session = 1

                    # Extract organization ID
                    org_match = re.search(r'organization (\d+)', line)
                    if org_match:
                        org_id = int(org_match.group(1))
                        if analysis[date][current_session]['org_id'] is None:
                            analysis[date][current_session]['org_id'] = org_id
                        if timestamp:
                            analysis[date][current_session]['session_start'] = timestamp

                    # Extract downloaded tenders
                    downloaded_match = re.search(r'Found (\d+) tender documents for (\d+) downloaded bids', line)
                    if downloaded_match:
                        analysis[date][current_session]['downloaded_tenders'] = int(downloaded_match.group(2))
                        analysis[date][current_session]['selected_tenders'] = int(downloaded_match.group(1))
                        if timestamp:
                            analysis[date][current_session]['session_end'] = timestamp

                    # Extract filtered out tenders
                    filtered_match = re.search(r'Tenders filtered out \(keyword score < 0\.1\): (\d+)', line)
                    if filtered_match:
                        analysis[date][current_session]['filtered_out_tenders'] = int(filtered_match.group(1))

                    # Extract analyzed tenders
                    analyzed_match = re.search(r'Tenders analyzed with API: (\d+)', line)
                    if analyzed_match:
                        analysis[date][current_session]['analyzed_tenders'] = int(analyzed_match.group(1))

                    # Extract ERROR lines (ignore warnings)
                    if ' - ERROR - ' in line:
                        error_msg = line.split(' - ERROR - ', 1)[1].strip() if ' - ERROR - ' in line else ''
                        error_entry = {
                            'timestamp': timestamp,
                            'line_number': line_num,
                            'message': error_msg[:200]  # Truncate long error messages
                        }
                        analysis[date][current_session]['errors'].append(error_entry)

                except Exception as e:
                    print(f"Error parsing line {line_num}: {e}")
                    continue

        return dict(analysis)

    except FileNotFoundError:
        print(f"Error: Log file '{log_file_path}' not found.")
        return {}
    except Exception as e:
        print(f"Error reading log file: {e}")
        return {}


def print_analysis(analysis):
    """
    Print the analysis results in a formatted way (grouped by date).
    """

    if not analysis:
        print("No analysis data available.")
        return

    print("\n" + "="*80)
    print("GEM TENDERS LOG ANALYSIS (Per-Date)")
    print("="*80 + "\n")

    for date in sorted(analysis.keys()):
        date_data = analysis[date]
        print(f"{'='*80}")
        print(f"DATE: {date}")
        print(f"{'='*80}")

        total_sessions_today = len(date_data)
        total_errors_today = sum(len(data['errors']) for data in date_data.values())
        total_downloaded_today = sum(data['downloaded_tenders'] for data in date_data.values())
        total_selected_today = sum(data['selected_tenders'] for data in date_data.values())
        total_analyzed_today = sum(data['analyzed_tenders'] for data in date_data.values())

        for session_id in sorted(date_data.keys()):
            data = date_data[session_id]
            print(f"\n  Session {session_id}:")
            if data['session_start']:
                print(f"  Start: {data['session_start']}")
            if data['session_end']:
                print(f"  End:   {data['session_end']}")
            print(f"  Organization ID:              {data['org_id']}")
            print(f"  Tenders Downloaded:          {data['downloaded_tenders']}")
            print(f"  Tenders Selected (after filter): {data['selected_tenders']}")
            print(f"  Tenders Filtered Out:        {data['filtered_out_tenders']}")
            print(f"  Tenders Analyzed:            {data['analyzed_tenders']}")
            print(f"  Errors:                      {len(data['errors'])}")

        # Print daily summary
        print(f"\n{'-'*80}")
        print(f"DAILY SUMMARY ({date})")
        print(f"{'-'*80}")
        print(f"Total Sessions:                {total_sessions_today}")
        print(f"Total Errors:                  {total_errors_today}")
        print(f"Total Tenders Downloaded:      {total_downloaded_today}")
        print(f"Total Tenders Selected:        {total_selected_today}")
        print(f"Total Tenders Analyzed:        {total_analyzed_today}")

        if total_downloaded_today > 0:
            selection_rate = (total_selected_today / total_downloaded_today) * 100
            print(f"Selection Rate:                {selection_rate:.2f}%")

        if total_selected_today > 0:
            analysis_rate = (total_analyzed_today / total_selected_today) * 100
            print(f"Analysis Rate:                 {analysis_rate:.2f}%")

        print()


def export_to_excel(analysis, output_dir='analysis_reports'):
    """
    Export analysis results to separate Excel files per date.
    Returns list of created files.
    """
    import os
    from openpyxl.utils import get_column_letter

    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)

    created_files = []

    for date in sorted(analysis.keys()):
        try:
            date_data = analysis[date]

            # Create filename with date
            date_str = date.replace('-', '_')
            filename = f"analysis_{date_str}.xlsx"
            filepath = os.path.join(output_dir, filename)

            workbook = Workbook()
            workbook.remove(workbook.active)  # Remove default sheet

            # Create Summary Sheet
            summary_sheet = workbook.create_sheet("Summary", 0)
            _create_summary_sheet(summary_sheet, date, date_data)

            # Create Details Sheet
            details_sheet = workbook.create_sheet("Details", 1)
            _create_details_sheet(details_sheet, date_data)

            # Create Errors Sheet if there are any errors
            error_count = sum(len(data['errors']) for data in date_data.values())
            if error_count > 0:
                errors_sheet = workbook.create_sheet("Errors", 2)
                _create_errors_sheet(errors_sheet, date_data)

            workbook.save(filepath)
            created_files.append(filepath)
            print(f"Exported analysis for {date} to: {filepath}")

        except Exception as e:
            print(f"Error exporting analysis for {date}: {e}")

    return created_files


def _create_summary_sheet(sheet, date, date_data):
    """Create summary sheet in Excel workbook."""
    from openpyxl.styles import Font, PatternFill, Alignment

    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    title_font = Font(bold=True, size=12)

    # Title
    sheet['A1'] = f"Analysis Summary - {date}"
    sheet['A1'].font = title_font
    sheet['A1'].fill = title_fill
    sheet.merge_cells('A1:B1')

    # Calculate totals
    total_sessions = len(date_data)
    total_errors = sum(len(data['errors']) for data in date_data.values())
    total_downloaded = sum(data['downloaded_tenders'] for data in date_data.values())
    total_selected = sum(data['selected_tenders'] for data in date_data.values())
    total_analyzed = sum(data['analyzed_tenders'] for data in date_data.values())

    row = 3

    # Add metrics
    metrics = [
        ("Total Sessions", total_sessions),
        ("Total Errors", total_errors),
        ("Total Tenders Downloaded", total_downloaded),
        ("Total Tenders Selected", total_selected),
        ("Total Tenders Analyzed", total_analyzed),
    ]

    for label, value in metrics:
        sheet[f'A{row}'] = label
        sheet[f'B{row}'] = value
        row += 1

    # Add percentages
    if total_downloaded > 0:
        selection_rate = (total_selected / total_downloaded) * 100
        sheet[f'A{row}'] = "Selection Rate"
        sheet[f'B{row}'] = f"{selection_rate:.2f}%"
        row += 1

    if total_selected > 0:
        analysis_rate = (total_analyzed / total_selected) * 100
        sheet[f'A{row}'] = "Analysis Rate"
        sheet[f'B{row}'] = f"{analysis_rate:.2f}%"

    # Auto-fit columns
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 20


def _create_details_sheet(sheet, date_data):
    """Create details sheet with per-session information."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ["Session", "Org ID", "Downloaded", "Selected", "Analyzed", "Filtered Out", "Errors", "Start Time", "End Time"]
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Data rows
    row = 2
    for session_id in sorted(date_data.keys()):
        data = date_data[session_id]
        sheet.cell(row=row, column=1).value = session_id
        sheet.cell(row=row, column=2).value = data['org_id']
        sheet.cell(row=row, column=3).value = data['downloaded_tenders']
        sheet.cell(row=row, column=4).value = data['selected_tenders']
        sheet.cell(row=row, column=5).value = data['analyzed_tenders']
        sheet.cell(row=row, column=6).value = data['filtered_out_tenders']
        sheet.cell(row=row, column=7).value = len(data['errors'])
        sheet.cell(row=row, column=8).value = data['session_start']
        sheet.cell(row=row, column=9).value = data['session_end']

        # Apply borders
        for col in range(1, len(headers) + 1):
            sheet.cell(row=row, column=col).border = border

        row += 1

    # Auto-fit columns
    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[chr(64 + col)].width = 15


def _create_errors_sheet(sheet, date_data):
    """Create errors sheet with all error details."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ["Session", "Timestamp", "Line Number", "Error Message"]
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Data rows
    row = 2
    for session_id in sorted(date_data.keys()):
        data = date_data[session_id]
        for error in data['errors']:
            sheet.cell(row=row, column=1).value = session_id
            sheet.cell(row=row, column=2).value = error['timestamp']
            sheet.cell(row=row, column=3).value = error['line_number']
            sheet.cell(row=row, column=4).value = error['message']

            # Apply borders and wrapping
            for col in range(1, len(headers) + 1):
                sheet.cell(row=row, column=col).border = border
                sheet.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

            row += 1

    # Auto-fit columns
    sheet.column_dimensions['A'].width = 10
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 12
    sheet.column_dimensions['D'].width = 50


def purge_old_sheets(output_dir='analysis_reports', days_to_keep=15):
    """
    Delete Excel files older than specified days.

    Args:
        output_dir: Directory containing Excel files
        days_to_keep: Number of days to keep files (default: 15)
    """
    import os
    import time

    if not os.path.exists(output_dir):
        print(f"Output directory '{output_dir}' not found.")
        return

    current_time = time.time()
    cutoff_time = current_time - (days_to_keep * 24 * 60 * 60)
    deleted_count = 0
    preserved_count = 0

    print(f"\nPurging Excel files older than {days_to_keep} days from '{output_dir}'...")

    for filename in os.listdir(output_dir):
        if filename.endswith('.xlsx') and filename.startswith('analysis_'):
            filepath = os.path.join(output_dir, filename)

            try:
                file_mtime = os.path.getmtime(filepath)

                if file_mtime < cutoff_time:
                    file_age_days = (current_time - file_mtime) / (24 * 60 * 60)
                    os.remove(filepath)
                    print(f"Deleted: {filename} (age: {file_age_days:.1f} days)")
                    deleted_count += 1
                else:
                    preserved_count += 1

            except Exception as e:
                print(f"Error deleting {filename}: {e}")

    print(f"\nPurge Summary:")
    print(f"  Files deleted: {deleted_count}")
    print(f"  Files preserved: {preserved_count}")


def main():
    """
    Main function to run the analysis.
    """
    import sys

    # Default log file path
    log_file = "gem_tenders.log"
    output_dir = "analysis_reports"
    days_to_keep = 15

    # Accept log file path from command line argument
    if len(sys.argv) > 1:
        log_file = sys.argv[1]

    if len(sys.argv) > 2:
        output_dir = sys.argv[2]

    if len(sys.argv) > 3:
        try:
            days_to_keep = int(sys.argv[3])
        except ValueError:
            print("Invalid days_to_keep argument. Using default (15).")
            days_to_keep = 15

    print(f"Analyzing log file: {log_file}")
    print(f"Output directory: {output_dir}")
    print(f"Days to keep: {days_to_keep}")

    # Parse the log file
    analysis = parse_log_file(log_file)

    if analysis:
        # Print the analysis
        print_analysis(analysis)

        # Export to Excel (separate files per date)
        created_files = export_to_excel(analysis, output_dir)
        print(f"\nCreated {len(created_files)} Excel report(s)")

        # Purge old sheets
        purge_old_sheets(output_dir, days_to_keep)
    else:
        print("No analysis data to export.")


if __name__ == "__main__":
    main()
